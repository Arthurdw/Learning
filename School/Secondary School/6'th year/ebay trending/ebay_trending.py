# -*- coding: utf-8 -*-
# MIT License

# Copyright (c) 2021 Arthurdw

# Permission is hereby granted, free of charge, to any person obtaining a copy
# of this software and associated documentation files (the "Software"), to deal
# in the Software without restriction, including without limitation the rights
# to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
# copies of the Software, and to permit persons to whom the Software is
# furnished to do so, subject to the following conditions:

# The above copyright notice and this permission notice shall be included in all
# copies or substantial portions of the Software.

# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
# FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
# AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
# LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
# OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
# SOFTWARE.

from datetime import datetime
from json import loads
from math import ceil
from typing import Optional, Dict, Union, List

from bs4 import BeautifulSoup
from openpyxl import Workbook
from requests import get


def get_all_items(limit: int):
    all_html = ""
    for i in range(ceil(limit / 24)):
        print(f"Scraped trending page with offset {i * 24}")
        txt = get(f"https://www.ebay.com/deals/spoke/ajax/listings?t=1622023095617&_ofs={i * 24}"
                  f"&category_path_seo=all&deal_type=trending").text
        all_html += loads(txt)["fulfillmentValue"]["listingsHtml"]

    soup = BeautifulSoup(all_html, "html.parser")

    return soup.findAll("div", class_="col")[:limit:]


def scrape(limit: int) -> List[Dict[Union[str, Dict[str, str]], str]]:
    data: List[Dict[Union[str, Dict[str, str]], str]] = []

    for item in get_all_items(limit):
        price = item.find("div", class_="dne-itemtile-price")
        original_price = item.find("div", class_="dne-itemtile-original-price")

        def get_original_price_prop(class_name: str) -> Optional[str]:
            el = original_price.find("span", class_=class_name) if original_price else None
            return el.getText() if el else None

        prce = {
            "type": "SOLD OUT",
            "price": None,
            "original": None,
            "discount": None
        } if price is None else {
            "type": price.find("meta")["content"],
            "price": price.getText(),
            "original": get_original_price_prop("itemtile-price-strikethrough"),
            "discount": get_original_price_prop("itemtile-price-bold")
        }

        data.append({
            "item": item.find("h3", class_="dne-itemtile-title").getText(),
            "url": item.find("a")["href"],
            "img": item.find("img")["src"],
            "price": prce
        })

    return data


def write_xlsx(name: str, data: List[Dict[Union[str, Dict[str, str]], str]]) -> None:
    if len(data) == 0:
        print("No data found.")
        return

    xlsx = Workbook()
    sheet_name = datetime.now().strftime("%d %b %Y - %I%p")
    sheet = xlsx.get_sheet_by_name(sheet_name) if sheet_name in xlsx.sheetnames else xlsx.create_sheet(sheet_name)

    keys = []
    for key in data[0].keys():
        if isinstance(data[0][key], dict):
            keys = [*keys, *data[0][key].keys()]
            continue
        keys.append(key)

    sheet.append(keys)

    for item in data:
        xlsx_data = []

        for value in item.values():
            if isinstance(value, dict):
                xlsx_data = [*xlsx_data, *value.values()]
            else:
                xlsx_data.append(value)

        sheet.append(xlsx_data)

    xlsx.save(name)


if __name__ == "__main__":
    print("How many records would you like to have?")
    try:
        item_count = int(input("Requested items (100): "))
    except ValueError:
        item_count = 100

    print("Started scraping Ebay trending page!")
    content = scrape(item_count)
    print("Started processing all data...")
    print("Scraped Ebay successfully!\n"
          "As what would you like to save the data? (.xlsx gets appended)")
    write_xlsx((input("File name (ebay_trending): ") or "ebay_trending") + ".xlsx", content)
    print("Successfully written data!")
